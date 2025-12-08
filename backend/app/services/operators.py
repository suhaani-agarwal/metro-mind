import json
import os
from datetime import datetime, date, timedelta
from typing import Dict, List, Any, Optional, Tuple
import logging
from pathlib import Path
import random
import sys
# Optional imports for enhanced scheduling using Layer 2 rotation
try:
    from app.services.layer2_service import run_layer2_service
    from app.utils.forecast import generate_rotation_schedule, get_station_timings, get_weather_forecast
except Exception:
    # If layer2/forecast or their dependencies are missing, we'll fallback to default generation
    run_layer2_service = None
    generate_rotation_schedule = None
    get_station_timings = None
    get_weather_forecast = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

logger = logging.getLogger(__name__)

# Shift configuration (8-hour duties). Currently only EARLY and LATE are used.
SHIFT_CONFIG = {
    "EARLY": {"start": "06:00", "end": "14:00"},
    "LATE": {"start": "14:00", "end": "22:00"},
}

SPARE_RATIO = 0.2  # 20% spare operators per shift
MAX_ACTIVE_OPERATORS_PER_DAY = 40  # 30-40 operators work daily (target upper bound)


def _parse_time(time_str: str) -> datetime:
    """Helper: parse HH:MM into a dummy-date datetime for arithmetic."""
    return datetime.strptime(time_str, "%H:%M")


def _format_time(dt: datetime) -> str:
    """Helper: format datetime (ignore date) back to HH:MM."""
    return dt.strftime("%H:%M")

class TrainOperatorService:
    """Service to generate train operator duty schedules"""
    
    def __init__(self):
        self.base_dir = Path(__file__).resolve().parent.parent.parent
        self.data_dir = os.path.join(self.base_dir, "data")
        self.storage_dir = os.path.join(self.base_dir, "storage")
        self.root_dir = self.base_dir.parent
        
        # Define operators (loaded from Excel if available, else synthetic)
        self.operator_db = self._create_operator_database()
        self.total_operators = len(self.operator_db)
        
        # Shift configurations
        self.shifts = {
            "EARLY": {"meal": "11:00-11:30"},
            "LATE": {"meal": "18:30-19:00"},
            "NIGHT": {"meal": "02:00-02:30"}
        }

        # Caches to avoid rerunning heavy Layer-2 / rotation for every operator
        self._rotation_cache: Dict[str, Dict[str, Any]] = {}
        self._optimized_trains_cache: Dict[str, List[Dict[str, Any]]] = {}
        self._train_shift_duties_cache: Dict[str, Dict[Tuple[str, str], Dict[str, Any]]] = {}
        self._assignments_cache: Dict[str, Dict[str, Dict[str, Any]]] = {}

        # In-memory overrides for manual leave / reassignment
        # Structure: { "YYYY-MM-DD": {"on_leave": set(operator_ids), "reassignments": [...] } }
        self._leave_overrides: Dict[str, Dict[str, Any]] = {}
    
    def _create_operator_database(self) -> Dict[str, Dict[str, Any]]:
        """
        Create operator database.
        Priority:
        1) Load from kochi_employee_database.xlsx if available
        2) Fallback to enriched built-in list (ensure ~60 operators)
        """
        # Try Excel-based operator list for realistic data
        try:
            if load_workbook:
                xlsx_path = self.root_dir / "kochi_employee_database.xlsx"
                if xlsx_path.exists():
                    wb = load_workbook(xlsx_path)
                    ws = wb.active
                    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

                    def col(*names):
                        for name in names:
                            if name in headers:
                                return headers[name]
                        return None

                    col_name = col("Name", "Employee Name", "EMPLOYEE NAME", "Operator Name")
                    col_emp_id = col("Employee ID", "Emp ID", "EMP ID", "EMPLOYEE ID", "Staff ID")
                    col_email = col("Email", "Email ID", "E-mail")
                    col_phone = col("Phone", "Mobile", "Phone Number", "Contact Number")
                    col_license = col("License Type", "License", "LIC TYPE", "License_Category")
                    col_exp = col("Experience (Years)", "Experience", "Years")

                    db: Dict[str, Dict[str, Any]] = {}
                    idx = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        name = row[col_name - 1] if col_name else None
                        emp_id = row[col_emp_id - 1] if col_emp_id else None
                        if not name or not emp_id:
                            continue
                        idx += 1
                        operator_id = f"KM-{idx:04d}-2025"
                        db[operator_id] = {
                            "name": str(name),
                            "phone": str(row[col_phone - 1]) if col_phone and row[col_phone - 1] else "",
                            "employee_id": str(emp_id),
                            "email": str(row[col_email - 1]) if col_email and row[col_email - 1] else "",
                            "experience_years": int(row[col_exp - 1]) if col_exp and isinstance(row[col_exp - 1], (int, float)) else random.randint(1, 12),
                            "license_type": str(row[col_license - 1]) if col_license and row[col_license - 1] else "A1",
                            "status": "active",
                        }
                    if idx >= 10:
                        logger.info(f"Loaded {idx} operators from Excel database")
                        return db
        except Exception as e:
            logger.warning(f"Failed to load operators from Excel: {e}")

        # Fallback: built-in synthetic database (60 operators)
        base_ops = [
            ("Rajesh Kumar", "SC/TO-0001"),
            ("Priya Sharma", "SC/TO-0002"),
            ("Amit Patel", "SC/TO-0003"),
            ("Sneha Nair", "SC/TO-0004"),
            ("Vikram Singh", "SC/TO-0005"),
            ("Ananya Das", "SC/TO-0006"),
            ("Rahul Menon", "SC/TO-0007"),
            ("Meera Krishnan", "SC/TO-0008"),
            ("Arjun Reddy", "SC/TO-0009"),
            ("Kiran Kumar", "SC/TO-0010"),
            ("Malika Verma", "SC/TO-0011"),
            ("Sanjay Rao", "SC/TO-0012"),
            ("Deepa Menon", "SC/TO-0013"),
            ("Nikhil Gupta", "SC/TO-0014"),
            ("Shruti Iyer", "SC/TO-0015"),
            ("Harish Nair", "SC/TO-0016"),
            ("Divya Pillai", "SC/TO-0017"),
            ("Manoj Kumar", "SC/TO-0018"),
            ("Rohit Sharma", "SC/TO-0019"),
            ("Pooja Singh", "SC/TO-0020"),
            ("Suresh Babu", "SC/TO-0021"),
            ("Farhan Ali", "SC/TO-0022"),
            ("Joseph Mathew", "SC/TO-0023"),
            ("Neha Thomas", "SC/TO-0024"),
            ("Vivek Mohan", "SC/TO-0025"),
            ("Anjali Rao", "SC/TO-0026"),
            ("Karthik Menon", "SC/TO-0027"),
            ("Lakshmi Prasad", "SC/TO-0028"),
            ("Ravi Shankar", "SC/TO-0029"),
            ("Meghna Nair", "SC/TO-0030"),
            ("Ajay Varma", "SC/TO-0031"),
            ("Geetha R", "SC/TO-0032"),
            ("Sameer Khan", "SC/TO-0033"),
            ("Ashwin Das", "SC/TO-0034"),
            ("Priyanka Pillai", "SC/TO-0035"),
            ("Sanjana R", "SC/TO-0036"),
            ("Arvind Krishnan", "SC/TO-0037"),
            ("Reena Joseph", "SC/TO-0038"),
            ("Mukesh Kumar", "SC/TO-0039"),
            ("Haritha Menon", "SC/TO-0040"),
            ("Gopalakrishnan", "SC/TO-0041"),
            ("Sunil Kumar", "SC/TO-0042"),
            ("Alka S", "SC/TO-0043"),
            ("Thomas Kurian", "SC/TO-0044"),
            ("Devika Menon", "SC/TO-0045"),
            ("Rajiv R", "SC/TO-0046"),
            ("Sneha Varma", "SC/TO-0047"),
            ("Aditya Rao", "SC/TO-0048"),
            ("Richa Sharma", "SC/TO-0049"),
            ("Vishnu Nair", "SC/TO-0050"),
            ("Anusha K", "SC/TO-0051"),
            ("Prakash Menon", "SC/TO-0052"),
            ("Sindhu R", "SC/TO-0053"),
            ("Deepak Gupta", "SC/TO-0054"),
            ("Monica Pillai", "SC/TO-0055"),
            ("Suraj Kumar", "SC/TO-0056"),
            ("Reshma N", "SC/TO-0057"),
            ("Kavya S", "SC/TO-0058"),
            ("Abhinav Rao", "SC/TO-0059"),
            ("Tarun Menon", "SC/TO-0060"),
        ]
        db: Dict[str, Dict[str, Any]] = {}
        for i in range(60):
            if i < len(base_ops):
                name, emp = base_ops[i]
            else:
                name = f"Operator {i+1:02d}"
                emp = f"SC/TO-{(i+1):04d}"
            operator_id = f"KM-{i+1:04d}-2025"
            db[operator_id] = {
                "name": name,
                "phone": f"+91-98{random.randint(100,999)}-{random.randint(10000,99999)}",
                "employee_id": emp,
                "email": f"{name.split()[0].lower()}.{name.split()[-1].lower()}@kmrl.co.in",
                "experience_years": random.randint(1, 12),
                "license_type": random.choice(["A1", "A2"]),
                "status": "active"
            }
        logger.info(f"Initialized synthetic operator database with {len(db)} operators")
        return db

    # ------------------------------------------------------------------
    # Core helper methods for using Layer-2 + rotation + unified.json
    # ------------------------------------------------------------------

    def _get_service_date_str(self, service_date: Optional[str]) -> str:
        if service_date:
            return service_date
        return date.today().isoformat()

    def _load_unified_config(self) -> Dict[str, Any]:
        unified_path = os.path.join(self.storage_dir, "unified.json")
        if not os.path.exists(unified_path):
            raise FileNotFoundError("Unified data (unified.json) not found. Run Layer 1 generation first.")
        with open(unified_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _ensure_rotation_data(self, service_date: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """
        Ensure we have Layer-2 optimized trains and full rotation schedule for the given date.
        Returns (rotation_schedule, optimized_trains).
        """
        if service_date in self._rotation_cache:
            return self._rotation_cache[service_date], self._optimized_trains_cache[service_date]

        if not run_layer2_service or not generate_rotation_schedule or not get_station_timings or not get_weather_forecast:
            raise RuntimeError("Layer 2 or forecast utilities are not available for operator scheduling")

        target_dt = date.fromisoformat(service_date)
        opt = run_layer2_service(service_day="weekday", service_date=target_dt, use_layer1_output=True)
        optimized_trains = opt.get("optimized_assignments") or []

        train_configs = self._load_unified_config()
        station_timings = get_station_timings()
        weather_data = get_weather_forecast(service_date)

        rotation_schedule = generate_rotation_schedule(
            scheduled_trains=optimized_trains,
            train_configs=train_configs,
            station_timings=station_timings,
            weather_data=weather_data,
            service_date=service_date
        )

        self._rotation_cache[service_date] = rotation_schedule
        self._optimized_trains_cache[service_date] = optimized_trains
        return rotation_schedule, optimized_trains

    def _build_master_timetable(self, rotation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Build master timetable: list of all trips (forward journeys) for the day."""
        timetable: List[Dict[str, Any]] = []

        for ts in rotation.get("train_schedules", []):
            train_id = ts.get("train_id")
            events = ts.get("station_events", [])
            rotations: Dict[int, List[Dict[str, Any]]] = {}
            for ev in events:
                if ev.get("direction") != "forward":
                    continue
                r = ev.get("rotation", 1)
                rotations.setdefault(r, []).append(ev)

            for rnum, evs in rotations.items():
                evs_sorted = sorted(evs, key=lambda e: e.get("cumulative_time", 0))
                origin_ev = evs_sorted[0]
                dest_ev = evs_sorted[-1]
                origin = origin_ev.get("station")
                destination = dest_ev.get("station")
                dep_time = origin_ev.get("scheduled_arrival")
                arr_time = dest_ev.get("scheduled_arrival")

                try:
                    d1 = _parse_time(dep_time)
                    d2 = _parse_time(arr_time)
                    if d2 < d1:
                        d2 += timedelta(days=1)
                    duration_min = int((d2 - d1).total_seconds() / 60)
                except Exception:
                    duration_min = 50

                timetable.append(
                    {
                        "trip_id": f"{train_id}-R{rnum}",
                        "train_id": train_id,
                        "origin": origin,
                        "destination": destination,
                        "scheduled_departure_time": dep_time,
                        "scheduled_arrival_time": arr_time,
                        "estimated_duration": duration_min,
                    }
                )

        timetable.sort(key=lambda t: (t["scheduled_departure_time"], t["train_id"]))
        return timetable

    def _build_train_roster(
        self,
        rotation: Dict[str, Any],
        trains_map: Dict[str, Dict[str, Any]],
        service_date: str
    ) -> Dict[str, Dict[str, Any]]:
        """Build train roster for the day from rotation + unified.json."""
        roster: Dict[str, Dict[str, Any]] = {}

        for ts in rotation.get("train_schedules", []):
            train_id = ts.get("train_id")
            first_trip = ts.get("first_departure")
            last_trip = ts.get("last_arrival")

            train_entry = trains_map.get(train_id) or next(
                (t for t in trains_map.values() if (t.get("train_id") or t.get("id")) == train_id),
                {}
            )
            job_cards = train_entry.get("job_cards", [])
            known_issues = [
                jc.get("description")
                for jc in job_cards
                if isinstance(jc, dict) and jc.get("criticality") in ("critical", "high", "medium")
            ]

            roster[train_id] = {
                "train_id": train_id,
                "date": service_date,
                "first_trip": first_trip,
                "last_trip": last_trip,
                "car_configuration": self._get_train_config(train_id),
                "known_issues": known_issues,
                "depot_location": train_entry.get("current_position", "Muttom Depot"),
            }

        return roster

    def _attach_predicted_delays(self, duty: Dict[str, Any]) -> None:
        """
        Attach predicted delays (when available) from predicted_delays.json
        generated by the /rotation/predictions endpoint.
        Only major delays (significant flag or >=5 min) are surfaced.
        """
        try:
            train_id = duty.get("train_id")
            trips = duty.get("trips") or []
            if not train_id or not trips:
                return

            predicted_path = os.path.join(self.storage_dir, "predicted_delays.json")
            if not os.path.exists(predicted_path):
                return

            with open(predicted_path, "r", encoding="utf-8") as f:
                prediction = json.load(f)

            train_schedules = prediction.get("train_schedules", [])
            ts = next((t for t in train_schedules if t.get("train_id") == train_id), None)
            if not ts:
                return

            events = ts.get("station_events", [])
            rotation_delay_map: Dict[int, List[Dict[str, Any]]] = {}
            for ev in events:
                rot = ev.get("rotation")
                if rot is None:
                    continue
                delay = float(ev.get("delay_minutes", 0) or 0)
                if ev.get("significant_delay") or delay >= 5.0:
                    rotation_delay_map.setdefault(int(rot), []).append(ev)

            major_delays_summary: List[str] = []
            for trip in trips:
                trip_no = trip.get("trip_number", "")
                if not trip_no.isdigit():
                    continue
                rot = int(trip_no)
                rot_events = rotation_delay_map.get(rot)
                if not rot_events:
                    continue

                # Pick the event with maximum delay for that rotation
                worst = max(rot_events, key=lambda e: float(e.get("delay_minutes", 0) or 0))
                delay_min = float(worst.get("delay_minutes", 0) or 0)
                station = worst.get("station", "")
                summary = f"Predicted delay {delay_min:.1f} min at {station}"

                trip["predicted_delay_minutes"] = delay_min
                trip["predicted_delay_summary"] = summary

                major_delays_summary.append(f"Trip {trip_no}: {summary}")

            if major_delays_summary:
                duty["major_predicted_delays"] = major_delays_summary
        except Exception as e:
            logger.warning(f"Failed to attach predicted delays to duty: {e}")

    def _build_train_shift_duties(self, service_date: str) -> None:
        """
        Pre-compute per-train, per-shift duty cores based on rotation schedule.
        These are later attached to specific operators.
        """
        if service_date in self._train_shift_duties_cache:
            return

        rotation, optimized_trains = self._ensure_rotation_data(service_date)
        train_schedules = {ts["train_id"]: ts for ts in rotation.get("train_schedules", [])}

        duties_for_date: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for train in optimized_trains:
            train_id = train.get("train_id")
            if not train_id or train_id not in train_schedules:
                continue
            ts = train_schedules[train_id]

            # Currently we only support EARLY and LATE shifts for revenue service
            for shift_name in ("EARLY", "LATE"):
                duty_core = self._build_day_shift_duty_core(train_id, ts, shift_name, service_date)
                if duty_core:
                    duties_for_date[(train_id, shift_name)] = duty_core

        self._train_shift_duties_cache[service_date] = duties_for_date

    def _build_day_shift_duty_core(
        self,
        train_id: str,
        train_schedule: Dict[str, Any],
        shift_name: str,
        service_date: str
    ) -> Optional[Dict[str, Any]]:
        """
        Build duty core for a given train+shift from rotation schedule.
        Applies:
        - Max 8h duty (within fixed shift window)
        - Continuous driving <= 2h (inserts 30min break)
        - Meal break ~ mid-shift avoiding peaks as much as possible
        - Sign-on 15min before first trip, sign-off 10min after last trip
        """
        shift_cfg = SHIFT_CONFIG[shift_name]
        shift_start = _parse_time(shift_cfg["start"])
        shift_end = _parse_time(shift_cfg["end"])
        if shift_end <= shift_start:
            shift_end = shift_end + timedelta(days=1)

        events = train_schedule.get("station_events", [])
        rotations: Dict[int, List[Dict[str, Any]]] = {}
        for ev in events:
            rotations.setdefault(ev.get("rotation", 1), []).append(ev)

        trips: List[Dict[str, Any]] = []
        first_trip_start: Optional[datetime] = None
        last_trip_end: Optional[datetime] = None

        for rnum, evs in sorted(rotations.items()):
            depart_ev = next(
                (e for e in evs if e.get("station") == "Aluva" and e.get("direction") == "forward"),
                None
            )
            arrive_ev = next(
                (
                    e
                    for e in evs
                    if e.get("station") in ("Tripunithura", "Tripunithura Terminal")
                    and e.get("direction") == "forward"
                ),
                None
            )
            if not depart_ev or not arrive_ev:
                depart_ev = evs[0]
                arrive_ev = evs[-1]

            dep_time = _parse_time(depart_ev.get("scheduled_arrival"))
            arr_time = _parse_time(arrive_ev.get("scheduled_arrival"))
            if arr_time < dep_time:
                arr_time += timedelta(days=1)

            # Filter by shift window
            if not (shift_start <= dep_time < shift_end):
                continue

            duration_min = int((arr_time - dep_time).total_seconds() / 60)
            hour = dep_time.hour
            if 7 <= hour <= 9 or 17 <= hour <= 19:
                pax_estimate = "Heavy"
                notes = "Peak"
            elif hour >= 22 or hour <= 5:
                pax_estimate = "Light"
                notes = "Night"
            else:
                pax_estimate = "Moderate"
                notes = "Normal"

            trips.append(
                {
                                "trip_number": str(rnum),
                    "route": "Aluva→Tripunithura",
                    "departure": _format_time(dep_time),
                    "arrival": _format_time(arr_time),
                                "duration": f"{duration_min} min",
                                "pax_estimate": pax_estimate,
                    "notes": notes,
                }
            )

            if first_trip_start is None or dep_time < first_trip_start:
                first_trip_start = dep_time
            if last_trip_end is None or arr_time > last_trip_end:
                last_trip_end = arr_time

        if not trips:
            return None

        trips_with_breaks, break_hours, meal_break_str = self._insert_breaks_and_meal(trips, shift_name)

        # Fixed 8-hour shift windows for EARLY and LATE duties
        shift_cfg = SHIFT_CONFIG[shift_name]
        sign_on = shift_cfg["start"]
        sign_off = shift_cfg["end"]
        total_hours = self._calculate_time_difference(sign_on, sign_off)

        driving_hours = sum(
            float(t.get("duration", "0").split()[0]) / 60.0
            for t in trips_with_breaks
            if t.get("trip_number") not in ("MB", "BRK")
        )

        train_config_str = self._get_train_config(train_id)

        duty_core: Dict[str, Any] = {
            "date": service_date,
            "duty_id_suffix": f"{train_id}-{shift_name}",
            "shift": f"{shift_name} TURN",
            "sign_on": sign_on,
            "sign_off": sign_off,
            "total_hours": round(total_hours, 2),
            "driving_hours": round(driving_hours, 2),
            "break_hours": round(break_hours, 2),
            "meal_break": meal_break_str,
            "standby_time": f"{self._subtract_time(sign_off, '00:30')}-{sign_off}",
            "train_id": train_id,
            "train_config": train_config_str,
            "trips": trips_with_breaks,
            "restrictions": self._get_operational_restrictions(),
            "emergency_contacts": self._get_emergency_contacts(),
            "pre_check_checklist": self._get_pre_check_checklist(),
            "reminders": self._get_mandatory_reminders(),
        }
        return duty_core

    def _build_night_shift_duty_core(self, train_id: str, service_date: str) -> Dict[str, Any]:
        """Night shift is primarily depot preparation / low-demand work."""
        shift_cfg = SHIFT_CONFIG["NIGHT"]
        shift_start = _parse_time(shift_cfg["start"])
        shift_end = _parse_time(shift_cfg["end"]) + timedelta(days=1)

        trips: List[Dict[str, Any]] = []
        current = shift_start + timedelta(minutes=15)

        prep_end = current + timedelta(minutes=45)
        trips.append(
            {
                "trip_number": "P1",
                "route": "Depot Prep & Systems Check",
                "departure": _format_time(current),
                "arrival": _format_time(prep_end),
                "duration": "45 min",
                "pax_estimate": "—",
                "notes": "Brake tests, CBTC verification, door checks",
            }
        )
        current = prep_end + timedelta(minutes=15)

        test_end = current + timedelta(minutes=60)
        trips.append(
            {
                "trip_number": "T1",
                "route": "Muttom Depot⇄Main Line Test",
                "departure": _format_time(current),
                "arrival": _format_time(test_end),
                "duration": "60 min",
                "pax_estimate": "—",
                "notes": "Low-speed test run & signal verification",
            }
        )
        current = test_end + timedelta(minutes=20)

        sb_end = min(current + timedelta(minutes=60), shift_end - timedelta(minutes=10))
        trips.append(
            {
                "trip_number": "SB",
                "route": "STANDBY @ Depot",
                "departure": _format_time(current),
                "arrival": _format_time(sb_end),
                "duration": f"{int((sb_end-current).total_seconds()/60)} min",
                "pax_estimate": "—",
                "notes": "Relief / incident response",
            }
        )

        trips_with_breaks, break_hours, meal_break_str = self._insert_breaks_and_meal(trips, "NIGHT")

        sign_on = _format_time(shift_start)
        sign_off = _format_time(shift_end)
        total_hours = self._calculate_time_difference(sign_on, sign_off)
        total_hours = min(total_hours, 10.0)

        driving_hours = sum(
            float(t.get("duration", "0").split()[0]) / 60.0
            for t in trips_with_breaks
            if t.get("trip_number") not in ("MB", "BRK")
        )

        train_config_str = self._get_train_config(train_id)

        duty_core: Dict[str, Any] = {
            "date": service_date,
            "duty_id_suffix": f"{train_id}-NIGHT",
            "shift": "NIGHT TURN",
            "sign_on": sign_on,
            "sign_off": sign_off,
            "total_hours": round(total_hours, 2),
            "driving_hours": round(driving_hours, 2),
            "break_hours": round(break_hours, 2),
            "meal_break": meal_break_str,
            "standby_time": f"{self._subtract_time(sign_off, '00:30')}-{sign_off}",
            "train_id": train_id,
            "train_config": train_config_str,
            "trips": trips_with_breaks,
            "restrictions": self._get_operational_restrictions(),
            "emergency_contacts": self._get_emergency_contacts(),
            "pre_check_checklist": self._get_pre_check_checklist(),
            "reminders": self._get_mandatory_reminders(),
        }
        return duty_core

    def _insert_breaks_and_meal(
        self, trips: List[Dict[str, Any]], shift_name: str
    ) -> Tuple[List[Dict[str, Any]], float, str]:
        """
        Insert mandatory short breaks & a 30-min meal break:
        - Max 2h continuous driving
        - Meal break ~ between 2nd-3rd trip and avoid peaks when possible
        """
        out: List[Dict[str, Any]] = []
        continuous_minutes = 0
        total_break_minutes = 0
        meal_break_str = self.shifts.get(shift_name, {}).get("meal", "")
        meal_inserted = False

        def is_peak(hour: int) -> bool:
            return 7 <= hour <= 10 or 16 <= hour <= 20

        for idx, trip in enumerate(trips):
            duration_min = int(trip.get("duration", "0").split()[0])
            continuous_minutes += duration_min
            out.append(trip)

            if (
                not meal_inserted
                and idx >= 1
                and continuous_minutes >= 90
            ):
                try:
                    arr_time = _parse_time(trip["arrival"])
                    if not is_peak(arr_time.hour):
                        mb_start = arr_time
                        mb_end = mb_start + timedelta(minutes=30)
                        out.append(
                            {
                                "trip_number": "MB",
                                "route": "MEAL BREAK",
                                "departure": _format_time(mb_start),
                                "arrival": _format_time(mb_end),
                                "duration": "30 min",
                                "pax_estimate": "—",
                                "notes": "Crew meal break",
                            }
                        )
                        continuous_minutes = 0
                        total_break_minutes += 30
                        meal_inserted = True
                        continue
                except Exception:
                    pass

            if continuous_minutes >= 120:
                try:
                    arr_time = _parse_time(trip["arrival"])
                    brk_start = arr_time
                    brk_end = brk_start + timedelta(minutes=20)
                    out.append(
                        {
                            "trip_number": "BRK",
                            "route": "TECHNICAL BREAK",
                            "departure": _format_time(brk_start),
                            "arrival": _format_time(brk_end),
                            "duration": "20 min",
                            "pax_estimate": "—",
                            "notes": "Mandatory rest after continuous driving",
                        }
                    )
                    continuous_minutes = 0
                    total_break_minutes += 20
                except Exception:
                    pass

        if not meal_inserted and meal_break_str:
            try:
                meal_start_str, meal_end_str = meal_break_str.split("-")
                out.append(
                    {
                        "trip_number": "MB",
                        "route": "MEAL BREAK",
                        "departure": meal_start_str,
                        "arrival": meal_end_str,
                        "duration": "30 min",
                        "pax_estimate": "—",
                        "notes": "Crew meal break",
                    }
                )
                total_break_minutes += 30
            except Exception:
                pass

        return out, total_break_minutes / 60.0, meal_break_str
    
    def get_all_operators(self) -> List[Dict[str, Any]]:
        """Get list of all operators"""
        operators = []
        for op_id, op_info in self.operator_db.items():
            operators.append({
                "id": op_id,
                "name": op_info["name"],
                "employee_id": op_info["employee_id"],
                "phone": op_info["phone"],
                "email": op_info["email"],
                "experience_years": op_info["experience_years"],
                "license_type": op_info["license_type"],
                "status": op_info["status"]
            })
        return sorted(operators, key=lambda x: x["id"])
    
    def get_operator_duty_summary(self, service_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Get duty summary for all operators for a given service date.

        This is driven by:
        - Layer 2 optimized trains (10 trains per day)
        - Rotation schedule (first/last trips)
        - Operator assignment rules (weekly off, spare pool, shifts)
        """
        service_date = self._get_service_date_str(service_date)

        self._build_train_shift_duties(service_date)
        assignments = self._compute_daily_assignments(service_date)

        summary = {
            "service_date": service_date,
            "generated_at": datetime.now().isoformat(),
            "total_operators": len(self.operator_db),
            "duty_assignments": []
        }

        for operator_id, op_info in self.operator_db.items():
            assign = assignments.get(operator_id, {})
            status = assign.get("status", "off")

            if status == "off":
                shift_label = "WEEKLY OFF"
                train_id = None
                sign_on = ""
                sign_off = ""
                total_hours = 0.0
                duty_id = f"OFF-{operator_id.split('-')[1].zfill(3)}"
            elif status == "on_leave":
                shift_label = "ON LEAVE"
                train_id = None
                sign_on = ""
                sign_off = ""
                total_hours = 0.0
                duty_id = f"LV-{operator_id.split('-')[1].zfill(3)}"
            elif status == "spare":
                shift_label = f"{assign.get('shift', '')} SPARE"
                train_id = "SPARE"
                sign_on = assign.get("sign_on", "")
                sign_off = assign.get("sign_off", "")
                total_hours = assign.get("total_hours", 8.0)
                duty_id = f"SP-{operator_id.split('-')[1].zfill(3)}"
            else:  # driver
                shift_label = assign.get("shift_label", "")
                train_id = assign.get("train_id")
                sign_on = assign.get("sign_on", "")
                sign_off = assign.get("sign_off", "")
                total_hours = assign.get("total_hours", 8.0)
                duty_id = f"D-{operator_id.split('-')[1].zfill(3)}"

            summary["duty_assignments"].append({
                "operator_id": operator_id,
                "operator_name": op_info["name"],
                "duty_id": duty_id,
                "shift": shift_label,
                "train_id": train_id,
                "sign_on": sign_on,
                "sign_off": sign_off,
                "total_hours": total_hours
            })

        # Enrich with master timetable, train roster, and operations bulletin for the day
        try:
            rotation, _ = self._ensure_rotation_data(service_date)
            unified = self._load_unified_config()
            trains_map = {t.get("id"): t for t in unified.get("trains", [])}

            summary["master_timetable"] = self._build_master_timetable(rotation)
            summary["train_roster"] = self._build_train_roster(rotation, trains_map, service_date)
            summary["operations_bulletin"] = {
                "date": service_date,
                "speed_restrictions": self._get_operational_restrictions(),
                "special_notes": [],
                "known_defects": [
                    {
                        "train_id": t_id,
                        "issues": roster_item.get("known_issues", [])
                    }
                    for t_id, roster_item in summary["train_roster"].items()
                    if roster_item.get("known_issues")
                ],
            }
        except Exception as e:
            logger.warning(f"Failed to build extended operator summary: {e}")

        return summary
    
    def generate_duty_schedule(self, operator_id: str, service_date: Optional[str] = None) -> Dict[str, Any]:
        """Generate duty schedule for a specific operator"""
        # Validate operator
        if operator_id not in self.operator_db:
            raise ValueError(f"Operator {operator_id} not found")
        
        service_date = self._get_service_date_str(service_date)
        operator_info = self.operator_db[operator_id]

        # Ensure we have shift/train assignments & precomputed train duties
        self._build_train_shift_duties(service_date)
        assignments = self._compute_daily_assignments(service_date)
        assignment = assignments.get(operator_id)

        if not assignment or assignment.get("status") in ("off", "on_leave"):
            status = assignment.get("status") if assignment else "off"
            is_leave = status == "on_leave"
            duty = {
                "date": service_date,
                "duty_id": (
                    f"LV-{operator_id.split('-')[1].zfill(3)}"
                    if is_leave
                    else f"OFF-{operator_id.split('-')[1].zfill(3)}"
                ),
                "shift": "ON LEAVE" if is_leave else "WEEKLY OFF",
                "operator_id": operator_id,
                "operator_name": operator_info["name"],
                "phone": operator_info["phone"],
                "employee_id": operator_info["employee_id"],
                "sign_on": "",
                "sign_off": "",
                "total_hours": 0.0,
                "driving_hours": 0.0,
                "break_hours": 0.0,
                "meal_break": "",
                "standby_time": "",
                "train_id": "",
                "train_config": "",
                "trips": [],
                "restrictions": [],
                "emergency_contacts": self._get_emergency_contacts(),
                "pre_check_checklist": [],
                "reminders": [
                    "On approved leave - no duty assigned (manual override)"
                    if is_leave
                    else "Weekly rest day - no duty assigned"
                ],
                "generated_at": datetime.now().isoformat(),
            }
            return duty

        status = assignment.get("status")
        shift_name = assignment.get("shift")
        duty_id_prefix = "D" if status == "driver" else "SP"

        if status == "spare":
            shift_cfg = SHIFT_CONFIG[shift_name]
            sign_on = shift_cfg["start"]
            sign_off = shift_cfg["end"]
            total_hours = self._calculate_time_difference(sign_on, sign_off)
            duty = {
                "date": service_date,
                "duty_id": f"{duty_id_prefix}-{operator_id.split('-')[1].zfill(3)}",
                "shift": f"{shift_name} SPARE",
                "operator_id": operator_id,
                "operator_name": operator_info["name"],
                "phone": operator_info["phone"],
                "employee_id": operator_info["employee_id"],
                "sign_on": sign_on,
                "sign_off": sign_off,
                "total_hours": total_hours,
                "driving_hours": 0.0,
                "break_hours": 2.0,
                "meal_break": self.shifts.get(shift_name, {}).get("meal", ""),
                "standby_time": f"{self._subtract_time(sign_off, '00:30')}-{sign_off}",
                "train_id": "SPARE",
                "train_config": "Spare duty - standby at Muttom Depot / Aluva",
                "trips": [
                    {
                        "trip_number": "SB",
                        "route": "SPARE / STANDBY",
                        "departure": sign_on,
                        "arrival": sign_off,
                        "duration": f"{int(total_hours*60)} min",
                        "pax_estimate": "—",
                        "notes": "Relief coverage, incident response, training"
                    }
                ],
                "restrictions": self._get_operational_restrictions(),
                "emergency_contacts": self._get_emergency_contacts(),
                "pre_check_checklist": self._get_pre_check_checklist(),
                "reminders": self._get_mandatory_reminders(),
                "generated_at": datetime.now().isoformat()
            }
            self._validate_duty(duty)
            return duty

        # Driver duty: attach operator information to precomputed train+shift duty core
        train_id = assignment.get("train_id")
        train_shift_key = (train_id, shift_name)
        duty_core = self._train_shift_duties_cache[service_date].get(train_shift_key)

        if not duty_core:
            # Fallback: generic pattern
            selected_shift = shift_name or "EARLY"
            basic_trips = self._generate_trips(selected_shift)
            sign_on = SHIFT_CONFIG[selected_shift]["start"]
            sign_off = SHIFT_CONFIG[selected_shift]["end"]
            total_hours = self._calculate_time_difference(sign_on, sign_off)
            duty = {
                "date": service_date,
                "duty_id": f"{duty_id_prefix}-{operator_id.split('-')[1].zfill(3)}",
                "shift": f"{selected_shift} TURN",
                "operator_id": operator_id,
                "operator_name": operator_info["name"],
                "phone": operator_info["phone"],
                "employee_id": operator_info["employee_id"],
                "sign_on": sign_on,
                "sign_off": sign_off,
                "total_hours": total_hours,
                "driving_hours": 4.5,
                "break_hours": 1.5,
                "meal_break": self.shifts[selected_shift]["meal"],
                "standby_time": f"{self._subtract_time(sign_off, '00:30')}-{sign_off}",
                "train_id": train_id,
                "train_config": self._get_train_config(train_id),
                "trips": basic_trips,
                "restrictions": self._get_operational_restrictions(),
                "emergency_contacts": self._get_emergency_contacts(),
                "pre_check_checklist": self._get_pre_check_checklist(),
                "reminders": self._get_mandatory_reminders(),
                "generated_at": datetime.now().isoformat(),
            }
            self._validate_duty(duty)
            return duty

        duty = {
            "date": duty_core["date"],
            "duty_id": f"{duty_id_prefix}-{operator_id.split('-')[1].zfill(3)}",
            "shift": duty_core["shift"],
            "operator_id": operator_id,
            "operator_name": operator_info["name"],
            "phone": operator_info["phone"],
            "employee_id": operator_info["employee_id"],
            "sign_on": duty_core["sign_on"],
            "sign_off": duty_core["sign_off"],
            "total_hours": duty_core["total_hours"],
            "driving_hours": duty_core["driving_hours"],
            "break_hours": duty_core["break_hours"],
            "meal_break": duty_core["meal_break"],
            "standby_time": duty_core["standby_time"],
            "train_id": duty_core["train_id"],
            "train_config": duty_core["train_config"],
            "trips": duty_core["trips"],
            "restrictions": duty_core["restrictions"],
            "emergency_contacts": duty_core["emergency_contacts"],
            "pre_check_checklist": duty_core["pre_check_checklist"],
            "reminders": duty_core["reminders"],
            "generated_at": datetime.now().isoformat()
        }

        # Enrich train_config with full train entry from unified.json for PDFs/UI
        try:
            unified = self._load_unified_config()
            trains = unified.get("trains", [])
            train_entry = next(
                (t for t in trains if (t.get("train_id") or t.get("id")) == duty["train_id"]),
                None
            )
            if train_entry:
                job_cards = train_entry.get("job_cards", [])
                known_issues = [
                    jc.get("description")
                    for jc in job_cards
                    if isinstance(jc, dict) and jc.get("criticality") in ("critical", "high", "medium")
                ]
                duty["train_config"] = {
                    "id": train_entry.get("id") or train_entry.get("train_id") or duty["train_id"],
                    "config_string": self._get_train_config(duty["train_id"]),
                    "status": train_entry.get("status"),
                    "job_cards": job_cards,
                    "known_issues": known_issues,
                    "depot_location": train_entry.get("current_position"),
                }
        except Exception:
            pass

        # Attach any major predicted delays (if available)
        self._attach_predicted_delays(duty)

        self._validate_duty(duty)
        return duty

    def _compute_daily_assignments(self, service_date: str) -> Dict[str, Dict[str, Any]]:
        """
        Compute per-operator assignments for a given date:
        - Weekly rest day (1 day/week per operator, rotating)
        - Exactly one shift per working operator (EARLY/LATE/NIGHT)
        - 20% spare pool per shift (rounded)
        - Each operating train gets at least one driver per EARLY & LATE shifts
        """
        if service_date in self._assignments_cache:
            return self._assignments_cache[service_date]

        svc_date = date.fromisoformat(service_date)
        weekday = svc_date.weekday()

        operator_ids = sorted(self.operator_db.keys())
        weekly_off: set = set()
        for idx, op_id in enumerate(operator_ids):
            # Simple rotating weekly off pattern based on operator index and weekday
            if idx % 7 == weekday:
                weekly_off.add(op_id)

        # Start with everyone except weekly off as candidate for work
        candidate_ops = [op for op in operator_ids if op not in weekly_off]

        # Limit total working operators to a realistic 30-40 range (cap at MAX_ACTIVE)
        active_ops = candidate_ops[:MAX_ACTIVE_OPERATORS_PER_DAY]
        # Remaining candidates beyond cap are treated as additional off
        extra_off = set(candidate_ops[MAX_ACTIVE_OPERATORS_PER_DAY:])

        active_count = len(active_ops)

        base_per_shift, rem = divmod(active_count, 2)
        shift_order = ["EARLY", "LATE"]
        shift_sizes: Dict[str, int] = {}
        for i, name in enumerate(shift_order):
            shift_sizes[name] = base_per_shift + (1 if i < rem else 0)

        self._build_train_shift_duties(service_date)
        train_shift_duties = self._train_shift_duties_cache.get(service_date, {})

        assignments: Dict[str, Dict[str, Any]] = {}
        # Everyone in weekly_off or extra_off has no duty today
        for op in weekly_off.union(extra_off):
            assignments[op] = {"status": "off"}

        active_index = 0
        for shift_name in shift_order:
            count = shift_sizes[shift_name]
            shift_ops = active_ops[active_index : active_index + count]
            active_index += count
            if not shift_ops:
                continue

            spare_count = max(1, int(round(len(shift_ops) * SPARE_RATIO))) if len(shift_ops) > 3 else 1
            if spare_count >= len(shift_ops):
                spare_count = max(1, len(shift_ops) - 1)

            driver_ops = shift_ops[:-spare_count] if spare_count > 0 else shift_ops
            spare_ops = shift_ops[-spare_count:] if spare_count > 0 else []

            available_combos = [
                (tid, s) for (tid, s) in train_shift_duties.keys() if s == shift_name
            ]
            if not available_combos:
                for op_id in shift_ops:
                    assignments[op_id] = {
                        "status": "spare",
                        "shift": shift_name,
                        "sign_on": SHIFT_CONFIG[shift_name]["start"],
                        "sign_off": SHIFT_CONFIG[shift_name]["end"],
                        "total_hours": self._calculate_time_difference(
                            SHIFT_CONFIG[shift_name]["start"], SHIFT_CONFIG[shift_name]["end"]
                        )
                    }
                continue

            num_combos = len(available_combos)
            for i, op_id in enumerate(driver_ops):
                combo = available_combos[i % num_combos]
                train_id, _ = combo
                duty_core = train_shift_duties.get(combo, {})
                assignments[op_id] = {
                    "status": "driver",
                    "shift": shift_name,
                    "shift_label": duty_core.get("shift", f"{shift_name} TURN"),
                    "train_id": train_id,
                    "sign_on": duty_core.get("sign_on", SHIFT_CONFIG[shift_name]["start"]),
                    "sign_off": duty_core.get("sign_off", SHIFT_CONFIG[shift_name]["end"]),
                    "total_hours": duty_core.get(
                        "total_hours",
                        self._calculate_time_difference(
                            SHIFT_CONFIG[shift_name]["start"], SHIFT_CONFIG[shift_name]["end"]
                        ),
                    ),
                }

            for op_id in spare_ops:
                assignments[op_id] = {
                    "status": "spare",
                    "shift": shift_name,
                    "sign_on": SHIFT_CONFIG[shift_name]["start"],
                    "sign_off": SHIFT_CONFIG[shift_name]["end"],
                    "total_hours": self._calculate_time_difference(
                        SHIFT_CONFIG[shift_name]["start"], SHIFT_CONFIG[shift_name]["end"]
                    ),
                }

        self._assignments_cache[service_date] = assignments
        return assignments

    def mark_operator_on_leave(self, operator_id: str, service_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Mark an operator as on leave for a given service date and
        reassign their duty to a spare operator in the same shift.

        - If the operator is already OFF/ON_LEAVE for that day, an error is raised.
        - A spare operator (status == 'spare') with the same shift is promoted to driver.
        - The original operator is marked with status 'on_leave'.
        """
        if operator_id not in self.operator_db:
            raise ValueError(f"Operator {operator_id} not found")

        # Default to "tomorrow" if not provided, since leave is usually for next day
        if service_date:
            target_date = service_date
        else:
            target_date = (date.today() + timedelta(days=1)).isoformat()

        # Ensure assignments and train duties are computed
        self._build_train_shift_duties(target_date)
        assignments = self._compute_daily_assignments(target_date)

        if operator_id not in assignments:
            raise ValueError(f"No duty assignment found for operator {operator_id} on {target_date}")

        current_assignment = assignments[operator_id]
        status = current_assignment.get("status")

        if status in ("off", "on_leave"):
            raise ValueError(
                f"Operator {operator_id} is already "
                f"{'on leave' if status == 'on_leave' else 'off'} on {target_date}"
            )

        shift_name = current_assignment.get("shift")
        if not shift_name:
            raise ValueError(f"Operator {operator_id} has no shift assigned on {target_date}")

        # Find a spare operator in the same shift
        spare_candidate_id: Optional[str] = None
        for op_id, assign in assignments.items():
            if op_id == operator_id:
                continue
            if assign.get("status") == "spare" and assign.get("shift") == shift_name:
                spare_candidate_id = op_id
                break

        if not spare_candidate_id:
            raise ValueError(
                f"No spare operator available in {shift_name} shift to cover leave for {operator_id}"
            )

        spare_assignment = assignments[spare_candidate_id]

        # Preserve previous state for debugging / UI
        previous_driver_assignment = current_assignment.copy()
        previous_spare_assignment = spare_assignment.copy()

        # Promote spare operator to driver, copying train duty from the original operator
        assignments[spare_candidate_id] = {
            **current_assignment,
            "status": "driver",
        }

        # Mark original operator as on leave (no train allocation)
        assignments[operator_id] = {
            "status": "on_leave",
            "shift": shift_name,
            "shift_label": "ON LEAVE",
            "train_id": None,
            "sign_on": "",
            "sign_off": "",
            "total_hours": 0.0,
        }

        # Update cache so subsequent calls (summary/duty) use the overridden assignments
        self._assignments_cache[target_date] = assignments

        # Track override in in-memory structure
        overrides_for_date = self._leave_overrides.setdefault(
            target_date, {"on_leave": set(), "reassignments": []}
        )
        overrides_for_date["on_leave"].add(operator_id)
        overrides_for_date["reassignments"].append(
            {
                "leave_operator_id": operator_id,
                "replacement_operator_id": spare_candidate_id,
                "created_at": datetime.now().isoformat(),
            }
        )

        # Build a fresh summary snapshot after override for frontend convenience
        updated_summary = self.get_operator_duty_summary(target_date)

        return {
            "success": True,
            "service_date": target_date,
            "message": (
                f"Operator {operator_id} marked on leave for {target_date}. "
                f"Spare operator {spare_candidate_id} reassigned to cover their duty."
            ),
            "leave_operator": {
                "operator_id": operator_id,
                "previous_assignment": previous_driver_assignment,
                "new_assignment": assignments[operator_id],
            },
            "replacement_operator": {
                "operator_id": spare_candidate_id,
                "previous_assignment": previous_spare_assignment,
                "new_assignment": assignments[spare_candidate_id],
            },
            "updated_summary": updated_summary,
        }
    
    def _generate_trips(self, shift_type: str) -> List[Dict[str, Any]]:
        """Generate trips based on shift type"""
        trips = []
        
        # Determine start time based on shift
        if shift_type == "EARLY":
            current_time = "06:15"
        elif shift_type == "LATE":
            current_time = "14:15"
        else:  # NIGHT
            current_time = "22:15"
        
        # Generate 5-6 trips with meal break
        trip_count = 5
        meal_break_index = 2
        
        for i in range(trip_count + 1):  # +1 for standby
            if i == meal_break_index:
                # Add meal break
                meal_start = self.shifts[shift_type]["meal"].split('-')[0]
                meal_end = self.shifts[shift_type]["meal"].split('-')[1]
                
                trips.append({
                    "trip_number": "MB",
                    "route": "MEAL BREAK",
                    "departure": meal_start,
                    "arrival": meal_end,
                    "duration": "30 min",
                    "pax_estimate": "—",
                    "notes": "Aluva Crew Rest Area"
                })
                
                current_time = self._add_time(meal_end, "00:10")
                continue
            
            if i == trip_count:
                # Add standby
                standby_start = self._subtract_time(trips[-1]["arrival"], "00:30")
                trips.append({
                    "trip_number": "SB",
                    "route": "STANDBY",
                    "departure": standby_start,
                    "arrival": trips[-1]["arrival"],
                    "duration": "30 min",
                    "pax_estimate": "—",
                    "notes": "Aluva"
                })
                break
            
            # Determine direction
            direction = "Aluva→Tripunithura" if i % 2 == 0 else "Tripunithura→Aluva"
            
            # Calculate arrival time
            arrival = self._add_time(current_time, "00:50")
            
            # Determine passenger estimate
            hour = int(current_time.split(':')[0])
            if 7 <= hour <= 9 or 17 <= hour <= 19:
                pax_estimate = "Heavy"
                notes = "Peak"
            elif hour >= 22 or hour <= 5:
                pax_estimate = "Light"
                notes = "Night"
            else:
                pax_estimate = "Moderate"
                notes = "Normal"
            
            trips.append({
                "trip_number": str(len(trips) + 1),
                "route": direction,
                "departure": current_time,
                "arrival": arrival,
                "duration": "50 min",
                "pax_estimate": pax_estimate,
                "notes": notes
            })
            
            # Add turnaround time
            current_time = self._add_time(arrival, "00:10")
        
        return trips
    
    def _get_train_config(self, train_id: str) -> str:
        """Get train configuration"""
        # Try to get from unified.json
        try:
            unified_path = os.path.join(self.storage_dir, "unified.json")
            if os.path.exists(unified_path):
                with open(unified_path, 'r') as f:
                    unified_data = json.load(f)
                trains = unified_data.get("trains", [])
                for train in trains:
                    tid = train.get("train_id") or train.get("id")
                    if tid == train_id:
                        capacity = train.get("capacity", 720)
                        manufacturer = train.get("manufacturer", "Alstom")
                        cars = train.get("cars", 6)
                        # Include key maintenance job cards in the configuration string
                        job_cards = train.get("job_cards", [])
                        critical_issues = [
                            jc.get("description")
                            for jc in job_cards
                            if isinstance(jc, dict) and jc.get("criticality") in ("critical", "high")
                        ]
                        issues_str = ""
                        if critical_issues:
                            issues_str = " | CRITICAL: " + "; ".join(critical_issues)
                        else:
                            medium_issues = [
                                jc.get("description")
                                for jc in job_cards
                                if isinstance(jc, dict) and jc.get("criticality") == "medium"
                            ]
                            if medium_issues:
                                issues_str = " | Issues: " + "; ".join(medium_issues)
                        return f"{manufacturer} {cars}-car | Capacity: {capacity} pax{issues_str}"
        except:
            pass
        
        # Default configuration
        try:
            train_num = int(''.join(filter(str.isdigit, train_id))) if any(char.isdigit() for char in train_id) else 7
        except:
            train_num = 7
            
        capacity = 720
        if train_num % 3 == 0:
            manufacturer = "Bombardier"
            cars = 4
            capacity = 480
        elif train_num % 3 == 1:
            manufacturer = "Alstom"
            cars = 6
            capacity = 720
        else:
            manufacturer = "Siemens"
            cars = 8
            capacity = 960
        
        return f"{manufacturer} {cars}-car | Capacity: {capacity} pax"
    
    def _get_operational_restrictions(self) -> List[Dict[str, Any]]:
        """Get operational restrictions"""
        return [
            {
                "type": "SPEED RESTRICTION",
                "description": "25 km/h Kalamassery-JLN Stadium (All trips)",
                "reason": "Track maintenance (till 18-Dec)",
                "action": "Obey speed board"
            },
            {
                "type": "SYSTEM STATUS",
                "description": "CBTC: Normal operation | Power: Normal | Track: Normal working"
            },
            {
                "type": "SPECIAL",
                "description": "None"
            }
        ]
    
    def _get_emergency_contacts(self) -> Dict[str, str]:
        """Get emergency contacts"""
        return {
            "emergency": "0484-2755-0000 (24/7)",
            "occ": "Radio Ch-1 / 275-5050 ext.1",
            "depot_control": "275-6000 ext.5",
            "aluva_sc": "Vikram +91-98XXXXXX",
            "tripunithura_sc": "Priya +91-98YYYYYY",
            "duty_incharge": "275-6000"
        }
    
    def _get_pre_check_checklist(self) -> List[str]:
        """Get pre-check checklist"""
        return [
            "Exterior: Body/couplers/wheels/lights",
            "Cab: Consoles A&B functional",
            "Emergency: Brakes/PA/alarms tested",
            "Defect Log: Signed & reviewed",
            "Train Ready: Confirmed with depot staff"
        ]
    
    def _get_mandatory_reminders(self) -> List[str]:
        """Get mandatory reminders"""
        return [
            "Arrive 15 min early",
            "Breath test required",
            "Read all sign-on notices",
            "Obey OCC instructions over printed schedule",
            "Report defects immediately",
            "NO mobile phone while driving/platform",
            "Log incidents at sign-off"
        ]
    
    def _validate_duty(self, duty: Dict[str, Any]):
        """Validate duty against operational rules"""
        violations = []
        
        if duty["total_hours"] > 10:
            violations.append(f"Total duty {duty['total_hours']}h exceeds 10h limit")
        
        if duty["driving_hours"] > 8:
            violations.append(f"Driving time {duty['driving_hours']}h exceeds 8h limit")
        
        if duty["break_hours"] < 1.5:
            violations.append(f"Break time {duty['break_hours']}h less than 1.5h minimum")
        
        if violations:
            logger.warning(f"Duty validation warnings: {violations}")
    
    def _calculate_time_difference(self, time1: str, time2: str) -> float:
        """Calculate time difference in hours"""
        try:
            t1 = datetime.strptime(time1, "%H:%M")
            t2 = datetime.strptime(time2, "%H:%M")
            
            if t2 < t1:
                # Cross midnight
                t2 = t2.replace(day=t2.day + 1)
            
            diff = (t2 - t1).seconds / 3600
            return round(diff, 2)
        except:
            return 8.0
    
    def _add_time(self, base_time: str, add_time: str) -> str:
        """Add time to base time"""
        try:
            base = datetime.strptime(base_time, "%H:%M")
            add_hours, add_minutes = map(int, add_time.split(":"))
            result = base + timedelta(hours=add_hours, minutes=add_minutes)
            return result.strftime("%H:%M")
        except:
            return base_time
    
    def _subtract_time(self, base_time: str, sub_time: str) -> str:
        """Subtract time from base time"""
        try:
            base = datetime.strptime(base_time, "%H:%M")
            sub_hours, sub_minutes = map(int, sub_time.split(":"))
            result = base - timedelta(hours=sub_hours, minutes=sub_minutes)
            return result.strftime("%H:%M")
        except:
            return base_time